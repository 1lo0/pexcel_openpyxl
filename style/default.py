#-*- coding:utf-8 -*-
# author:donttouchkeyboard@gmail.com
# software: PyCharm


from openpyxl.styles import PatternFill, Border, Side, Alignment, Protection, Font,NamedStyle


class ExcelFontStyle:
    """
    excel字体风格，背景
    """
    __default_style__ = NamedStyle('default')
    __default_style__.border = Border(left=Side(border_style='thin'),
                                  right=Side(border_style='thin'),
                                  top=Side(border_style='thin'),
                                  bottom=Side(border_style='thin'))
    __default_style__.font = Font(size=12)
    __default_style__.alignment = Alignment(horizontal='center', vertical='center', )

    # header_style
    __header_style__ = NamedStyle("header_style")
    __header_style__.font = Font(bold=True, size=18)

    __header_style__.fill = PatternFill(fill_type='solid',
                                        start_color='00C0C0C0',
                                        end_color='00C0C0C0',)

    __header_style__.border = Border(left=Side(border_style='thin'),
                                  right=Side(border_style='thin'),
                                  top=Side(border_style='thin'),
                                  bottom=Side(border_style='thin'))

    __header_style__.alignment = Alignment(horizontal='center', vertical='center', )
    __header_style__.number_format = 'General'

    #openpyxl默认样式
    __openpyxl_default_style__ = NamedStyle("openpyxl_style")
    __openpyxl_default_style__.font = Font(name='Calibri',
                size=11,
                bold=False,
                italic=False,
                vertAlign=None,
                underline='none',
                strike=False,
                color='FF000000')
    __openpyxl_default_style__.fill = PatternFill(fill_type=None,
                       start_color='FFFFFFFF',
                       end_color='FF000000')
    __openpyxl_default_style__.border = Border(left=Side(border_style=None,
                              color='FF000000'),
                    right=Side(border_style=None,
                               color='FF000000'),
                    top=Side(border_style=None,
                             color='FF000000'),
                    bottom=Side(border_style=None,
                                color='FF000000'),
                    diagonal=Side(border_style=None,
                                  color='FF000000'),
                    diagonal_direction=0,
                    outline=Side(border_style=None,
                                 color='FF000000'),
                    vertical=Side(border_style=None,
                                  color='FF000000'),
                    horizontal=Side(border_style=None,
                                    color='FF000000')
                    )
    __openpyxl_default_style__.alignment = Alignment(horizontal='general',
                          vertical='bottom',
                          text_rotation=0,
                          wrap_text=False,
                          shrink_to_fit=False,
                          indent=0)
    __openpyxl_default_style__.number_format = 'General'
    __openpyxl_default_style__.protection = Protection(locked=True,
                            hidden=False)


    @classmethod
    def get_default_style(self):
        """
         设置单元格格式，其中字体及背景为默认
         #TODO 表格样式和筛选
        :return:
        """
        return self.__default_style__

    @classmethod
    def get_openpyxl_default_style(self):
        return self.__openpyxl_default_style__

    @classmethod
    def get_header_style(self):
        return self.__header_style__

